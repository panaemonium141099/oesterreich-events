"""Bundle the 11 audit CSVs in data/exports/ into one .xlsx workbook,
one sheet per bundesland + a summary sheet up front."""
import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# openpyxl rejects ASCII control chars (0x00-0x1F except tab/lf/cr) and
# the unicode replacement char block — strip them so scraper junk like
# "�Respekt im Netz�" doesn't blow up the writer.
ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]')

def clean(v):
    if isinstance(v, str):
        return ILLEGAL.sub('', v)
    return v

EXPORTS = Path(__file__).resolve().parents[2] / 'data' / 'exports'
OUT = EXPORTS / 'events_audit.xlsx'

ORDER = [
    'districts_summary',
    'events_clean',           # Tab 2: rows the map actually renders
    'events_burgenland',
    'events_kaernten',
    'events_niederoesterreich',
    'events_oberoesterreich',
    'events_salzburg',
    'events_steiermark',
    'events_tirol',
    'events_vorarlberg',
    'events_wien',
    'events__no_bundesland',
]

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='2C3E50')
BODY_FONT = Font(name='Arial', size=10)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

wb = Workbook()
wb.remove(wb.active)

for stem in ORDER:
    csv_path = EXPORTS / f'{stem}.csv'
    if not csv_path.exists():
        print(f'  skip {stem} (file missing)')
        continue

    sheet_name = stem.replace('events_', '').replace('_', ' ')[:31] or 'sheet'
    ws = wb.create_sheet(title=sheet_name)

    with csv_path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append([clean(h) for h in header])
        for row in reader:
            ws.append([clean(c) for c in row])

    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_LEFT

    # Sensible widths capped to keep the file usable
    for col_idx, name in enumerate(header, start=1):
        letter = get_column_letter(col_idx)
        if name == 'title':
            ws.column_dimensions[letter].width = 50
        elif name in ('id', 'source_url'):
            ws.column_dimensions[letter].width = 36
        elif name in ('district', 'district_pre_normalize', 'location_name', 'category'):
            ws.column_dimensions[letter].width = 22
        elif name == 'start_date':
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 14

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    print(f'  added {sheet_name}: {ws.max_row - 1} rows, {ws.max_column} cols')

wb.save(OUT)
print(f'\nWrote {OUT}')
print(f'Size: {OUT.stat().st_size / 1024 / 1024:.1f} MB')
